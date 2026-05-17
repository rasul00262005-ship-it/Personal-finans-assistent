import tkinter as tk
from tkinter import ttk, messagebox
import pandas as pd
import os
import matplotlib.pyplot as plt
import seaborn as sns
from datetime import datetime, timedelta
import json
from tkcalendar import DateEntry
import warnings
warnings.filterwarnings('ignore')

# ========== CONFIGURATION ==========
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
FILE_NAME = os.path.join(BASE_DIR, "finance_data.csv")
BUDGET_FILE = os.path.join(BASE_DIR, "budget.json")

# ========== CATEGORIES ==========
CATEGORIES = {
    "Income": ["Salary", "Freelance", "Gift", "Investment", "Cashback", "Other"],
    "Expense": ["Food", "Transport", "Housing", "Entertainment", "Health", 
                "Clothing", "Education", "Communication", "Other"]
}

# ========== DATA LOADING ==========
def load_data():
    if os.path.exists(FILE_NAME):
        df = pd.read_csv(FILE_NAME)
        if 'Date' in df.columns:
            df['Date'] = pd.to_datetime(df['Date'], errors='coerce')
        if 'Amount' in df.columns:
            df['Amount'] = pd.to_numeric(df['Amount'], errors='coerce').fillna(0)
        df = df.fillna("")
        return df
    else:
        return pd.DataFrame(columns=["Date", "Type", "Category", "Amount", "Note"])

def save_data(df):
    df_save = df.copy()
    df_save = df_save.fillna("")
    df_save.to_csv(FILE_NAME, index=False, encoding='utf-8-sig')

# ========== BUDGET LOADING ==========
def load_budget():
    if os.path.exists(BUDGET_FILE):
        try:
            with open(BUDGET_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        except:
            return {cat: 0 for cat in CATEGORIES["Expense"]}
    return {cat: 0 for cat in CATEGORIES["Expense"]}

def save_budget(budget):
    with open(BUDGET_FILE, 'w', encoding='utf-8') as f:
        json.dump(budget, f, ensure_ascii=False, indent=2)

# ========== MAIN APPLICATION ==========
class FinanceApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Personal Finance Assistant")
        self.root.geometry("900x700")
        self.root.configure(bg='#f5f5f5')
        
        self.df = load_data()
        self.budget = load_budget()
        
        self.create_widgets()
        
        self.update_balance()
        self.refresh_records()
        self.update_settings_info()
        self.update_quick_stats()
        
        self.root.protocol("WM_DELETE_WINDOW", self.on_closing)
    
    def create_widgets(self):
        # Header
        header_frame = tk.Frame(self.root, bg='#3498db', height=70)
        header_frame.pack(fill=tk.X)
        header_frame.pack_propagate(False)
        
        tk.Label(header_frame, text="💰 PERSONAL FINANCE ASSISTANT", 
                font=('Arial', 20, 'bold'), bg='#3498db', fg='white').pack(expand=True)
        
        # Main container
        self.notebook = ttk.Notebook(self.root)
        self.notebook.pack(fill=tk.BOTH, expand=True, padx=15, pady=15)
        
        self.create_add_tab()
        self.create_records_tab()
        self.create_stats_tab()
        self.create_settings_tab()
    
    def create_add_tab(self):
        tab = tk.Frame(self.notebook, bg='#f5f5f5')
        self.notebook.add(tab, text="➕ ADD RECORD")
        
        # Center container
        center_frame = tk.Frame(tab, bg='#f5f5f5')
        center_frame.pack(expand=True)
        
        # Card frame
        card = tk.Frame(center_frame, bg='white', bd=0, relief=tk.FLAT)
        card.pack(padx=40, pady=30, fill=tk.BOTH, expand=True)
        
        # Title
        tk.Label(card, text="New Transaction", font=('Arial', 18, 'bold'), 
                bg='white', fg='#2c3e50').pack(pady=(20, 30))
        
        # Form fields
        form_frame = tk.Frame(card, bg='white')
        form_frame.pack(pady=10)
        
        # Type
        tk.Label(form_frame, text="Type:", font=('Arial', 12), bg='white', 
                fg='#555', width=12, anchor='w').grid(row=0, column=0, pady=8, sticky='w')
        
        self.type_var = tk.StringVar(value="Expense")
        type_frame = tk.Frame(form_frame, bg='white')
        type_frame.grid(row=0, column=1, pady=8, sticky='w')
        
        tk.Radiobutton(type_frame, text="💸 Expense", variable=self.type_var, 
                      value="Expense", command=self.update_categories,
                      bg='white', fg='#e74c3c', font=('Arial', 11)).pack(side=tk.LEFT, padx=10)
        tk.Radiobutton(type_frame, text="💰 Income", variable=self.type_var, 
                      value="Income", command=self.update_categories,
                      bg='white', fg='#27ae60', font=('Arial', 11)).pack(side=tk.LEFT, padx=10)
        
        # Category
        tk.Label(form_frame, text="Category:", font=('Arial', 12), bg='white', 
                fg='#555', width=12, anchor='w').grid(row=1, column=0, pady=8, sticky='w')
        
        self.category_var = tk.StringVar()
        self.category_combo = ttk.Combobox(form_frame, textvariable=self.category_var, 
                                           state="readonly", width=30, font=('Arial', 11))
        self.category_combo.grid(row=1, column=1, pady=8)
        
        # Amount
        tk.Label(form_frame, text="Amount:", font=('Arial', 12), bg='white', 
                fg='#555', width=12, anchor='w').grid(row=2, column=0, pady=8, sticky='w')
        
        self.amount_entry = tk.Entry(form_frame, font=('Arial', 14), width=32, 
                                     bd=1, relief=tk.SOLID, highlightthickness=0)
        self.amount_entry.grid(row=2, column=1, pady=8)
        
        # Date
        tk.Label(form_frame, text="Date:", font=('Arial', 12), bg='white', 
                fg='#555', width=12, anchor='w').grid(row=3, column=0, pady=8, sticky='w')
        
        self.date_entry = DateEntry(form_frame, width=28, background='#3498db',
                                    foreground='white', borderwidth=0, date_pattern='yyyy-mm-dd')
        self.date_entry.grid(row=3, column=1, pady=8)
        
        # Note
        tk.Label(form_frame, text="Note:", font=('Arial', 12), bg='white', 
                fg='#555', width=12, anchor='w').grid(row=4, column=0, pady=8, sticky='w')
        
        self.note_text = tk.Text(form_frame, height=3, width=32, font=('Arial', 11),
                                  bd=1, relief=tk.SOLID)
        self.note_text.grid(row=4, column=1, pady=8)
        
        # Buttons
        btn_frame = tk.Frame(card, bg='white')
        btn_frame.pack(pady=30)
        
        tk.Button(btn_frame, text="💾 SAVE RECORD", command=self.add_record,
                 bg='#27ae60', fg='white', font=('Arial', 11, 'bold'),
                 padx=25, pady=8, bd=0, cursor='hand2').pack(side=tk.LEFT, padx=10)
        
        tk.Button(btn_frame, text="🗑 CLEAR", command=self.clear_form,
                 bg='#95a5a6', fg='white', font=('Arial', 11, 'bold'),
                 padx=25, pady=8, bd=0, cursor='hand2').pack(side=tk.LEFT, padx=10)
        
        # Balance display
        balance_frame = tk.Frame(card, bg='#ecf0f1', bd=0, relief=tk.FLAT)
        balance_frame.pack(fill=tk.X, pady=20, padx=20)
        
        self.balance_label = tk.Label(balance_frame, text="", font=('Arial', 22, 'bold'),
                                       bg='#ecf0f1', fg='#2c3e50')
        self.balance_label.pack(pady=10)
        
        self.income_label = tk.Label(balance_frame, text="", font=('Arial', 11),
                                      bg='#ecf0f1', fg='#27ae60')
        self.income_label.pack()
        
        self.expense_label = tk.Label(balance_frame, text="", font=('Arial', 11),
                                       bg='#ecf0f1', fg='#e74c3c')
        self.expense_label.pack(pady=(0, 10))
        
        self.update_categories()
    
    def create_records_tab(self):
        tab = tk.Frame(self.notebook, bg='#f5f5f5')
        self.notebook.add(tab, text="📋 ALL RECORDS")
        
        # Filters
        filter_frame = tk.Frame(tab, bg='white', bd=0, relief=tk.FLAT)
        filter_frame.pack(fill=tk.X, padx=15, pady=10)
        
        tk.Label(filter_frame, text="Filters:", font=('Arial', 11, 'bold'), 
                bg='white', fg='#555').pack(side=tk.LEFT, padx=10)
        
        self.filter_type = ttk.Combobox(filter_frame, values=["All", "Income", "Expense"], 
                                        state="readonly", width=12)
        self.filter_type.set("All")
        self.filter_type.pack(side=tk.LEFT, padx=5)
        
        all_cats = ["All"] + CATEGORIES["Expense"] + CATEGORIES["Income"]
        self.filter_category = ttk.Combobox(filter_frame, values=all_cats,
                                            state="readonly", width=15)
        self.filter_category.set("All")
        self.filter_category.pack(side=tk.LEFT, padx=5)
        
        self.filter_date_from = DateEntry(filter_frame, width=12, date_pattern='yyyy-mm-dd')
        self.filter_date_from.pack(side=tk.LEFT, padx=5)
        
        self.filter_date_to = DateEntry(filter_frame, width=12, date_pattern='yyyy-mm-dd')
        self.filter_date_to.set_date(datetime.now())
        self.filter_date_to.pack(side=tk.LEFT, padx=5)
        
        tk.Button(filter_frame, text="Apply", command=self.apply_filter,
                 bg='#3498db', fg='white', font=('Arial', 10),
                 padx=15, pady=3, bd=0, cursor='hand2').pack(side=tk.LEFT, padx=5)
        
        tk.Button(filter_frame, text="Reset", command=self.reset_filter,
                 bg='#95a5a6', fg='white', font=('Arial', 10),
                 padx=15, pady=3, bd=0, cursor='hand2').pack(side=tk.LEFT, padx=5)
        
        # Table
        self.create_records_table(tab)
        
        # Control buttons
        control_frame = tk.Frame(tab, bg='#f5f5f5')
        control_frame.pack(pady=10)
        
        tk.Button(control_frame, text="✏ Edit", command=self.edit_record,
                 bg='#f39c12', fg='white', font=('Arial', 10, 'bold'),
                 padx=20, pady=5, bd=0, cursor='hand2').pack(side=tk.LEFT, padx=5)
        
        tk.Button(control_frame, text="🗑 Delete", command=self.delete_record,
                 bg='#e74c3c', fg='white', font=('Arial', 10, 'bold'),
                 padx=20, pady=5, bd=0, cursor='hand2').pack(side=tk.LEFT, padx=5)
        
        self.record_count_label = tk.Label(tab, text="", font=('Arial', 10), 
                                           bg='#f5f5f5', fg='#777')
        self.record_count_label.pack(pady=5)
    
    def create_records_table(self, parent):
        table_frame = tk.Frame(parent, bg='#f5f5f5')
        table_frame.pack(fill=tk.BOTH, expand=True, padx=15, pady=10)
        
        columns = ("ID", "Date", "Type", "Category", "Amount", "Note")
        self.tree = ttk.Treeview(table_frame, columns=columns, show="headings", height=15)
        
        self.tree.heading("ID", text="ID")
        self.tree.heading("Date", text="Date")
        self.tree.heading("Type", text="Type")
        self.tree.heading("Category", text="Category")
        self.tree.heading("Amount", text="Amount")
        self.tree.heading("Note", text="Note")
        
        self.tree.column("ID", width=50, anchor='center')
        self.tree.column("Date", width=100, anchor='center')
        self.tree.column("Type", width=80, anchor='center')
        self.tree.column("Category", width=130)
        self.tree.column("Amount", width=100, anchor='e')
        self.tree.column("Note", width=300)
        
        v_scroll = ttk.Scrollbar(table_frame, orient=tk.VERTICAL, command=self.tree.yview)
        h_scroll = ttk.Scrollbar(table_frame, orient=tk.HORIZONTAL, command=self.tree.xview)
        self.tree.configure(yscrollcommand=v_scroll.set, xscrollcommand=h_scroll.set)
        
        self.tree.grid(row=0, column=0, sticky="nsew")
        v_scroll.grid(row=0, column=1, sticky="ns")
        h_scroll.grid(row=1, column=0, sticky="ew")
        
        table_frame.grid_rowconfigure(0, weight=1)
        table_frame.grid_columnconfigure(0, weight=1)
    
    def create_stats_tab(self):
        tab = tk.Frame(self.notebook, bg='#f5f5f5')
        self.notebook.add(tab, text="📊 STATISTICS")
        
        # Center container
        center_frame = tk.Frame(tab, bg='#f5f5f5')
        center_frame.pack(expand=True, fill=tk.BOTH, padx=20, pady=20)
        
        # Chart buttons
        btn_container = tk.Frame(center_frame, bg='#f5f5f5')
        btn_container.pack(pady=10)
        
        buttons = [
            ("📊 Income & Expenses", self.show_bar_chart, '#3498db'),
            ("🥧 Expense Structure", self.show_pie_chart, '#e74c3c'),
            ("📉 Monthly Trend", self.show_monthly_trend, '#27ae60'),
            ("🔥 Weekly Expenses", self.show_heatmap, '#f39c12'),
            ("📋 Detailed Stats", self.show_detailed_stats, '#9b59b6')
        ]
        
        for text, cmd, color in buttons:
            tk.Button(btn_container, text=text, command=cmd,
                     bg=color, fg='white', font=('Arial', 10, 'bold'),
                     padx=20, pady=8, bd=0, cursor='hand2', width=20).pack(side=tk.LEFT, padx=8)
        
        # Quick stats
        stats_frame = tk.Frame(center_frame, bg='white', bd=0, relief=tk.FLAT)
        stats_frame.pack(fill=tk.BOTH, expand=True, pady=20)
        
        tk.Label(stats_frame, text="Quick Statistics", font=('Arial', 16, 'bold'),
                bg='white', fg='#2c3e50').pack(pady=(20, 10))
        
        self.quick_stats_text = tk.Text(stats_frame, height=10, font=('Courier', 11),
                                        bg='white', fg='#2c3e50', bd=0, wrap=tk.WORD)
        self.quick_stats_text.pack(fill=tk.BOTH, expand=True, padx=20, pady=10)
    
    def create_settings_tab(self):
        tab = tk.Frame(self.notebook, bg='#f5f5f5')
        self.notebook.add(tab, text="⚙ SETTINGS")
        
        # Center container
        center_frame = tk.Frame(tab, bg='#f5f5f5')
        center_frame.pack(expand=True, fill=tk.BOTH, padx=30, pady=30)
        
        # Export section
        export_frame = tk.Frame(center_frame, bg='white', bd=0, relief=tk.FLAT)
        export_frame.pack(fill=tk.X, pady=10)
        
        tk.Label(export_frame, text="📤 Export Data", font=('Arial', 14, 'bold'),
                bg='white', fg='#2c3e50').pack(pady=(15, 10))
        
        btn_export_frame = tk.Frame(export_frame, bg='white')
        btn_export_frame.pack(pady=10)
        
        for text, cmd in [("CSV", self.export_csv), ("Excel", self.export_excel), ("PDF", self.export_pdf)]:
            tk.Button(btn_export_frame, text=text, command=cmd,
                     bg='#27ae60', fg='white', font=('Arial', 10, 'bold'),
                     padx=25, pady=6, bd=0, cursor='hand2', width=10).pack(side=tk.LEFT, padx=10)
        
        # Data management
        data_frame = tk.Frame(center_frame, bg='white', bd=0, relief=tk.FLAT)
        data_frame.pack(fill=tk.X, pady=10)
        
        tk.Label(data_frame, text="⚠ Data Management", font=('Arial', 14, 'bold'),
                bg='white', fg='#2c3e50').pack(pady=(15, 10))
        
        btn_data_frame = tk.Frame(data_frame, bg='white')
        btn_data_frame.pack(pady=10)
        
        tk.Button(btn_data_frame, text="🗑 Clear All Data", command=self.clear_all_data,
                 bg='#e74c3c', fg='white', font=('Arial', 10, 'bold'),
                 padx=25, pady=6, bd=0, cursor='hand2', width=15).pack(side=tk.LEFT, padx=10)
        
        tk.Button(btn_data_frame, text="💾 Create Backup", command=self.backup_data,
                 bg='#3498db', fg='white', font=('Arial', 10, 'bold'),
                 padx=25, pady=6, bd=0, cursor='hand2', width=15).pack(side=tk.LEFT, padx=10)
        
        # Info section (without frame)
        info_frame = tk.Frame(center_frame, bg='white', bd=0, relief=tk.FLAT)
        info_frame.pack(fill=tk.BOTH, expand=True, pady=10)
        
        tk.Label(info_frame, text="Application Info", font=('Arial', 14, 'bold'),
                bg='white', fg='#2c3e50').pack(pady=(15, 10))
        
        self.settings_info_text = tk.Text(info_frame, height=8, font=('Courier', 11),
                                          bg='white', fg='#2c3e50', bd=0, wrap=tk.WORD)
        self.settings_info_text.pack(fill=tk.BOTH, expand=True, padx=20, pady=10)
    
    # ========== CORE FUNCTIONS ==========
    
    def update_categories(self):
        self.category_combo['values'] = CATEGORIES[self.type_var.get()]
        self.category_combo.set(CATEGORIES[self.type_var.get()][0])
    
    def clear_form(self):
        self.amount_entry.delete(0, tk.END)
        self.note_text.delete("1.0", tk.END)
        self.date_entry.set_date(datetime.now())
    
    def get_total_income(self):
        if self.df.empty:
            return 0.0
        inc = self.df[self.df['Type'] == 'Income']
        return float(inc['Amount'].sum()) if not inc.empty else 0.0
    
    def get_total_expense(self):
        if self.df.empty:
            return 0.0
        exp = self.df[self.df['Type'] == 'Expense']
        return float(exp['Amount'].sum()) if not exp.empty else 0.0
    
    def update_balance(self):
        inc = self.get_total_income()
        exp = self.get_total_expense()
        self.balance_label.config(text=f"💰 {inc - exp:,.2f}")
        self.income_label.config(text=f"📈 Income: {inc:,.2f}")
        self.expense_label.config(text=f"📉 Expenses: {exp:,.2f}")
    
    def update_settings_info(self):
        total = len(self.df)
        inc = self.get_total_income()
        exp = self.get_total_expense()
        
        info = f"""
Total Records:  {total}
Total Income:   {inc:,.2f}
Total Expense:  {exp:,.2f}
Net Balance:    {inc - exp:,.2f}
"""
        self.settings_info_text.config(state=tk.NORMAL)
        self.settings_info_text.delete("1.0", tk.END)
        self.settings_info_text.insert("1.0", info)
        self.settings_info_text.config(state=tk.DISABLED)
    
    def update_quick_stats(self):
        if self.df.empty:
            self.quick_stats_text.config(state=tk.NORMAL)
            self.quick_stats_text.delete("1.0", tk.END)
            self.quick_stats_text.insert("1.0", "No data available")
            self.quick_stats_text.config(state=tk.DISABLED)
            return
        
        inc = self.get_total_income()
        exp = self.get_total_expense()
        
        exp_data = self.df[self.df['Type'] == 'Expense']
        top_cat = "N/A"
        top_amt = 0
        if not exp_data.empty:
            top = exp_data.groupby('Category')['Amount'].sum().sort_values(ascending=False)
            if len(top) > 0:
                top_cat = top.index[0]
                top_amt = top.iloc[0]
        
        stats = f"""
Total Income:     {inc:>12,.2f}
Total Expense:    {exp:>12,.2f}
Net Balance:      {inc - exp:>12,.2f}

Top Expense:      {top_cat:<15} {top_amt:>10,.2f}
Total Records:    {len(self.df):>12}
"""
        self.quick_stats_text.config(state=tk.NORMAL)
        self.quick_stats_text.delete("1.0", tk.END)
        self.quick_stats_text.insert("1.0", stats)
        self.quick_stats_text.config(state=tk.DISABLED)
    
    def get_filtered_data(self):
        if self.df.empty:
            return self.df
        filtered = self.df.copy()
        if hasattr(self, 'filter_type') and self.filter_type.get() != "All":
            filtered = filtered[filtered['Type'] == self.filter_type.get()]
        if hasattr(self, 'filter_category') and self.filter_category.get() != "All":
            filtered = filtered[filtered['Category'] == self.filter_category.get()]
        if hasattr(self, 'filter_date_from'):
            from_date = self.filter_date_from.get_date()
            filtered = filtered[pd.to_datetime(filtered['Date']) >= pd.to_datetime(from_date)]
        if hasattr(self, 'filter_date_to'):
            to_date = self.filter_date_to.get_date()
            filtered = filtered[pd.to_datetime(filtered['Date']) <= pd.to_datetime(to_date)]
        return filtered
    
    def refresh_records(self):
        for item in self.tree.get_children():
            self.tree.delete(item)
        
        filtered = self.get_filtered_data()
        for idx, row in filtered.iterrows():
            date_str = row['Date'].strftime('%Y-%m-%d') if hasattr(row['Date'], 'strftime') else str(row['Date'])
            note_val = str(row.get('Note', '')) if row.get('Note', '') else ''
            self.tree.insert("", tk.END, values=(
                idx, date_str, str(row['Type']), str(row['Category']),
                f"{float(row['Amount']):,.2f}", note_val[:50]
            ))
        
        self.record_count_label.config(text=f"Showing {len(filtered)} of {len(self.df)} records")
        self.update_quick_stats()
        self.update_settings_info()
        self.update_balance()
    
    def apply_filter(self):
        self.refresh_records()
    
    def reset_filter(self):
        if hasattr(self, 'filter_type'):
            self.filter_type.set("All")
        if hasattr(self, 'filter_category'):
            self.filter_category.set("All")
        if hasattr(self, 'filter_date_from'):
            self.filter_date_from.set_date(datetime.now() - timedelta(days=30))
        if hasattr(self, 'filter_date_to'):
            self.filter_date_to.set_date(datetime.now())
        self.refresh_records()
    
    def add_record(self):
        try:
            amount = float(self.amount_entry.get())
            if amount <= 0:
                raise ValueError
        except:
            messagebox.showerror("Error", "Please enter a valid amount!")
            return
        
        if not self.category_var.get():
            messagebox.showerror("Error", "Please select a category!")
            return
        
        new_rec = {
            "Date": self.date_entry.get_date(),
            "Type": self.type_var.get(),
            "Category": self.category_var.get(),
            "Amount": amount,
            "Note": self.note_text.get("1.0", tk.END).strip()
        }
        
        self.df = pd.concat([self.df, pd.DataFrame([new_rec])], ignore_index=True)
        self.df['Amount'] = pd.to_numeric(self.df['Amount'], errors='coerce').fillna(0)
        save_data(self.df)
        
        self.update_balance()
        self.refresh_records()
        self.clear_form()
        messagebox.showinfo("Success", "Record added successfully!")
    
    def delete_record(self):
        selected = self.tree.selection()
        if not selected:
            messagebox.showwarning("Warning", "Please select a record to delete!")
            return
        
        if messagebox.askyesno("Confirm", "Delete this record?"):
            idx = int(self.tree.item(selected[0])['values'][0])
            self.df = self.df.drop(idx).reset_index(drop=True)
            save_data(self.df)
            self.refresh_records()
            self.update_balance()
            messagebox.showinfo("Success", "Record deleted!")
    
    def edit_record(self):
        selected = self.tree.selection()
        if not selected:
            messagebox.showwarning("Warning", "Please select a record to edit!")
            return
        
        idx = int(self.tree.item(selected[0])['values'][0])
        rec = self.df.loc[idx]
        
        win = tk.Toplevel(self.root)
        win.title("Edit Record")
        win.geometry("450x550")
        win.configure(bg='#f5f5f5')
        
        f = tk.Frame(win, bg='white', padx=30, pady=20)
        f.pack(fill=tk.BOTH, expand=True)
        
        tk.Label(f, text="Edit Transaction", font=('Arial', 16, 'bold'),
                bg='white', fg='#2c3e50').pack(pady=(0, 20))
        
        # Type
        tk.Label(f, text="Type:", font=('Arial', 11), bg='white', anchor='w').pack(fill=tk.X, pady=5)
        tv = tk.StringVar(value=rec['Type'])
        tf = tk.Frame(f, bg='white')
        tf.pack(fill=tk.X, pady=5)
        tk.Radiobutton(tf, text="Expense", variable=tv, value="Expense", bg='white').pack(side=tk.LEFT, padx=10)
        tk.Radiobutton(tf, text="Income", variable=tv, value="Income", bg='white').pack(side=tk.LEFT, padx=10)
        
        # Category
        tk.Label(f, text="Category:", font=('Arial', 11), bg='white', anchor='w').pack(fill=tk.X, pady=5)
        cat = ttk.Combobox(f, values=CATEGORIES["Expense"] + CATEGORIES["Income"], state="readonly")
        cat.set(rec['Category'])
        cat.pack(fill=tk.X, pady=5)
        
        # Amount
        tk.Label(f, text="Amount:", font=('Arial', 11), bg='white', anchor='w').pack(fill=tk.X, pady=5)
        amt = tk.Entry(f, font=('Arial', 12))
        amt.insert(0, str(rec['Amount']))
        amt.pack(fill=tk.X, pady=5)
        
        # Date
        tk.Label(f, text="Date:", font=('Arial', 11), bg='white', anchor='w').pack(fill=tk.X, pady=5)
        dt = DateEntry(f, date_pattern='yyyy-mm-dd')
        dt.set_date(rec['Date'])
        dt.pack(fill=tk.X, pady=5)
        
        # Note
        tk.Label(f, text="Note:", font=('Arial', 11), bg='white', anchor='w').pack(fill=tk.X, pady=5)
        nt = tk.Text(f, height=3)
        nt.insert("1.0", rec.get('Note', ''))
        nt.pack(fill=tk.X, pady=5)
        
        def save():
            try:
                a = float(amt.get())
                if a <= 0: raise ValueError
            except:
                messagebox.showerror("Error", "Invalid amount!")
                return
            
            self.df.loc[idx, 'Type'] = tv.get()
            self.df.loc[idx, 'Category'] = cat.get()
            self.df.loc[idx, 'Amount'] = a
            self.df.loc[idx, 'Date'] = dt.get_date()
            self.df.loc[idx, 'Note'] = nt.get("1.0", tk.END).strip()
            save_data(self.df)
            self.refresh_records()
            self.update_balance()
            win.destroy()
            messagebox.showinfo("Success", "Record updated!")
        
        tk.Button(f, text="SAVE CHANGES", command=save,
                 bg='#27ae60', fg='white', font=('Arial', 11, 'bold'),
                 padx=20, pady=8, bd=0, cursor='hand2').pack(pady=20)
    
    # ========== CHARTS ==========
    
    def show_bar_chart(self):
        if self.df.empty:
            messagebox.showwarning("No Data", "Add records first!")
            return
        summary = self.df.groupby(["Category", "Type"])["Amount"].sum().reset_index()
        pivot = summary.pivot(index="Category", columns="Type", values="Amount").fillna(0)
        pivot.plot(kind='bar', color=['#e74c3c', '#2ecc71'])
        plt.title("Income vs Expenses by Category")
        plt.xticks(rotation=45)
        plt.tight_layout()
        plt.show()
    
    def show_pie_chart(self):
        if self.df.empty:
            messagebox.showwarning("No Data", "Add records first!")
            return
        exp = self.df[self.df['Type'] == 'Expense'].groupby('Category')['Amount'].sum()
        if exp.empty:
            messagebox.showwarning("No Data", "No expenses!")
            return
        plt.figure(figsize=(10, 8))
        plt.pie(exp.values, labels=exp.index, autopct='%1.1f%%')
        plt.title("Expense Structure")
        plt.show()
    
    def show_monthly_trend(self):
        if self.df.empty:
            messagebox.showwarning("No Data", "Add records first!")
            return
        df2 = self.df.copy()
        df2['Month'] = pd.to_datetime(df2['Date']).dt.to_period('M')
        monthly = df2.groupby(['Month', 'Type'])['Amount'].sum().unstack().fillna(0)
        monthly.plot(kind='line', marker='o')
        plt.title("Monthly Trend")
        plt.xticks(rotation=45)
        plt.tight_layout()
        plt.show()
    
    def show_heatmap(self):
        if self.df.empty:
            messagebox.showwarning("No Data", "Add records first!")
            return
        exp = self.df[self.df['Type'] == 'Expense'].copy()
        if exp.empty:
            messagebox.showwarning("No Data", "No expenses!")
            return
        exp['Day'] = pd.to_datetime(exp['Date']).dt.dayofweek
        avg = exp.groupby('Day')['Amount'].mean()
        days = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun']
        plt.figure(figsize=(10, 6))
        plt.bar(days, avg.values, color='skyblue')
        plt.title("Average Expenses by Day")
        plt.show()
    
    def show_detailed_stats(self):
        if self.df.empty:
            messagebox.showwarning("No Data", "Add records first!")
            return
        
        win = tk.Toplevel(self.root)
        win.title("Detailed Statistics")
        win.geometry("600x500")
        win.configure(bg='#f5f5f5')
        
        txt = tk.Text(win, font=('Courier', 11), bg='white', fg='#2c3e50', padx=20, pady=20)
        txt.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)
        
        inc = self.get_total_income()
        exp = self.get_total_expense()
        
        stats = f"""
DETAILED FINANCIAL STATISTICS
{'='*50}

Total Income:     {inc:>15,.2f}
Total Expense:    {exp:>15,.2f}
Net Balance:      {inc - exp:>15,.2f}

Income Transactions:  {len(self.df[self.df['Type'] == 'Income'])}
Expense Transactions: {len(self.df[self.df['Type'] == 'Expense'])}
Total Transactions:   {len(self.df)}

{'='*50}
"""
        txt.insert("1.0", stats)
        txt.config(state=tk.DISABLED)
    
    # ========== EXPORT ==========
    
    def export_csv(self):
        if self.df.empty:
            messagebox.showwarning("No Data", "Nothing to export!")
            return
        fname = f"export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        self.df.to_csv(os.path.join(BASE_DIR, fname), index=False)
        messagebox.showinfo("Success", f"Saved: {fname}")
    
    def export_excel(self):
        try:
            import openpyxl
            if self.df.empty:
                messagebox.showwarning("No Data", "Nothing to export!")
                return
            fname = f"export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            self.df.to_excel(os.path.join(BASE_DIR, fname), index=False)
            messagebox.showinfo("Success", f"Saved: {fname}")
        except ImportError:
            messagebox.showerror("Error", "Install openpyxl: pip install openpyxl")
    
    def export_pdf(self):
        try:
            from fpdf import FPDF
            import urllib.request
            
            if self.df.empty:
                messagebox.showwarning("No Data", "Nothing to export!")
                return
            
            # Папка для шрифтов
            font_dir = os.path.join(BASE_DIR, "fonts")
            os.makedirs(font_dir, exist_ok=True)
            font_path = os.path.join(font_dir, "DejaVuSans.ttf")
            
            # Скачиваем шрифт если его нет
            if not os.path.exists(font_path):
                try:
                    url = "https://github.com/dejavu-fonts/dejavu-fonts/raw/master/ttf/DejaVuSans.ttf"
                    urllib.request.urlretrieve(url, font_path)
                except:
                    pass
            
            # Создаём PDF с поддержкой Unicode
            pdf = FPDF()
            
            if os.path.exists(font_path):
                pdf.add_font('DejaVu', '', font_path, uni=True)
                pdf.set_font('DejaVu', '', 10)
            else:
                pdf.set_font('helvetica', '', 10)
            
            pdf.add_page()
            
            # Заголовок
            if os.path.exists(font_path):
                pdf.set_font('DejaVu', 'B', 16)
            else:
                pdf.set_font('helvetica', 'B', 16)
            pdf.cell(0, 10, 'Financial Report', 0, 1, 'C')
            
            if os.path.exists(font_path):
                pdf.set_font('DejaVu', '', 10)
            else:
                pdf.set_font('helvetica', '', 10)
            pdf.cell(0, 8, f'Date: {datetime.now().strftime("%Y-%m-%d")}', 0, 1, 'C')
            pdf.ln(5)
            
            # Сводка
            if os.path.exists(font_path):
                pdf.set_font('DejaVu', 'B', 12)
            else:
                pdf.set_font('helvetica', 'B', 12)
            pdf.cell(0, 10, 'Summary:', 0, 1)
            
            if os.path.exists(font_path):
                pdf.set_font('DejaVu', '', 10)
            else:
                pdf.set_font('helvetica', '', 10)
            pdf.cell(0, 8, f'Total Income: {self.get_total_income():,.2f}', 0, 1)
            pdf.cell(0, 8, f'Total Expense: {self.get_total_expense():,.2f}', 0, 1)
            pdf.cell(0, 8, f'Balance: {self.get_total_income() - self.get_total_expense():,.2f}', 0, 1)
            pdf.ln(5)
            
            # Заголовки таблицы
            if os.path.exists(font_path):
                pdf.set_font('DejaVu', 'B', 9)
            else:
                pdf.set_font('helvetica', 'B', 9)
            
            pdf.cell(35, 8, 'Date', 1)
            pdf.cell(35, 8, 'Type', 1)
            pdf.cell(50, 8, 'Category', 1)
            pdf.cell(30, 8, 'Amount', 1)
            pdf.cell(0, 8, 'Note', 1, 1)
            
            # Данные
            if os.path.exists(font_path):
                pdf.set_font('DejaVu', '', 8)
            else:
                pdf.set_font('helvetica', '', 8)
            
            for _, row in self.df.iterrows():
                d = row['Date'].strftime('%Y-%m-%d') if hasattr(row['Date'], 'strftime') else str(row['Date'])
                note = str(row.get('Note', ''))[:30]
                
                pdf.cell(35, 7, d, 1)
                pdf.cell(35, 7, row['Type'], 1)
                pdf.cell(50, 7, row['Category'], 1)
                pdf.cell(30, 7, f"{row['Amount']:,.2f}", 1, 0, 'R')
                pdf.cell(0, 7, note, 1, 1)
            
            fname = f"report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
            pdf.output(os.path.join(BASE_DIR, fname))
            messagebox.showinfo("Success", f"Saved: {fname}")
            
        except ImportError:
            messagebox.showerror("Error", "Install fpdf2: pip install fpdf2")
        except Exception as e:
            messagebox.showerror("Error", f"PDF export failed: {str(e)}")
    
    def clear_all_data(self):
        if messagebox.askyesno("Confirm", "Delete ALL data? This cannot be undone!"):
            self.df = pd.DataFrame(columns=["Date", "Type", "Category", "Amount", "Note"])
            save_data(self.df)
            self.refresh_records()
            self.update_balance()
            self.update_quick_stats()
            self.update_settings_info()
            messagebox.showinfo("Success", "All data deleted!")
    
    def backup_data(self):
        if self.df.empty:
            messagebox.showwarning("No Data", "Nothing to backup!")
            return
        fname = f"backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        self.df.to_csv(os.path.join(BASE_DIR, fname), index=False)
        messagebox.showinfo("Success", f"Backup saved: {fname}")
    
    def on_closing(self):
        save_data(self.df)
        self.root.destroy()

# ========== RUN ==========
if __name__ == "__main__":
    root = tk.Tk()
    app = FinanceApp(root)
    root.mainloop()